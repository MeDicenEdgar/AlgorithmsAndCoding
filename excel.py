import openpyxl
from openpyxl.cell import cell
from openpyxl.workbook import workbook
from openpyxl.workbook.workbook import Workbook
lista = [1, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 541, 582, 583, 584, 585, 586, 587, 588, 589, 590, 591, 12441, 12442, 12443, 12444, 12477, 12478, 12479, 12480]
excel = openpyxl.load_workbook("Athena.xlsx")
pagina=excel.active
resultadosXLS=Workbook()
paginaXLS=resultadosXLS.active

contadorFila=1
contadorColumna=1
for i in lista:
    fila = pagina[i]
    for n in fila:
        n.value
        NuevaCelda=paginaXLS.cell(row=contadorFila, column=contadorColumna)
        NuevaCelda.value=n.value 
        print("{}/{}".format(contadorColumna,contadorFila))
        contadorColumna+=1
    contadorColumna=1    
    contadorFila+=1


resultadosXLS.save('resultados.xlsx')

"""for i in lista:
    fila = pagina[i]
    paginaXLS[contadorFila]=fila
    contadorFila+=1
    print(contadorFila)"""
